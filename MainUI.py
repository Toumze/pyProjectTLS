import sys

import xlwt
import openpyxl
from PyQt5 import QtCore, QtWidgets
from PyQt5.QtWidgets import *
import shark
import time


class MainUI(QWidget):
    def __init__(self):
        super(MainUI, self).__init__(None)
        self.tshark_path = None
        self.interface = None
        self.bpf_filter = None
        self.display_filter = None
        self.file_rode = None
        self.choice_function = None
        self.shark = None
        self.window_width = 800
        self.window_height = 600
        self.setWindowTitle("TLS指纹分析")

        self.resize(self.window_width, self.window_height)
        self.setMinimumWidth(self.window_width)  # 设置最小窗口大小
        self.setMinimumHeight(self.window_height)
        print(self.width(), self.height())

        # 选择 tshark.exe
        self.label_tshark = QLabel(self)
        self.label_tshark.setText("tshark：")
        self.lineEdit_tshark = QLineEdit(self)
        self.lineEdit_tshark.setText("D:\\software\\Wireshark\\tshark.exe")
        self.button_open_tshark = QPushButton(self)
        self.button_open_tshark.setText("选择")
        self.button_open_tshark.clicked.connect(self.choose_tshark)

        # 选择活动接口读取
        self.button_live_capture = QPushButton(self)
        self.button_live_capture.setText("接口读取")
        self.button_live_capture.clicked.connect(self.choose_live_capture)
        self.comboBox_live_interface = QComboBox(self)

        # 选择文件读取
        self.button_file_capture = QPushButton(self)
        self.button_file_capture.setText("文件读取")
        self.button_file_capture.clicked.connect(self.choose_file_capture)
        self.lineEdit_file_capture = QLineEdit(self)

        # 过滤规则
        self.label_display_filter = QLabel(self)
        self.label_display_filter.setText("过滤规则:")
        self.lineEdit_display_filter = QLineEdit(self)

        # 分割线

        # 确认按钮
        self.button_begin = QPushButton(self)
        self.button_begin.setText("开始")
        self.button_begin.clicked.connect(self.begin)

        # 测试按钮
        self.button_test = QPushButton(self)
        self.button_test.setText("测试")
        self.button_test.clicked.connect(self.test_reprint)

        # 显示区域
        self.tableWidget_overall = QTableWidget(self)
        self.tableWidget = QTableWidget(self)

        # 导出数据
        self.button_save_excel = QPushButton(self)
        self.button_save_excel.setText("导出")
        self.button_save_excel.clicked.connect(self.save_excel)

    def resizeEvent(self, a0) -> None:
        self.window_height = a0.size().height()
        self.window_width = a0.size().width()
        self.reprint()

    def reprint(self):
        change = self.window_width * 1.0 / 800

        # 位置 tshark.exe
        self.label_tshark.setGeometry(QtCore.QRect(int(change * 40), 30, 75, 23))
        self.lineEdit_tshark.setGeometry(QtCore.QRect(int(change * 40) + 65, 30, 180, 23))
        self.button_open_tshark.setGeometry(QtCore.QRect(int(change * 40) + 215 + 40, 30, 75, 23))

        # 位置 活动接口读取
        self.button_live_capture.setGeometry(QtCore.QRect(int(change * 40), 100, 80, 25))
        self.comboBox_live_interface.setGeometry(QtCore.QRect(int(change * 40) + 90, 100, 200, 25))

        # 位置 选择文件读取
        self.button_file_capture.setGeometry(QtCore.QRect(int(change * 400), 100, 80, 25))
        self.lineEdit_file_capture.setGeometry(QtCore.QRect(int(change * 400) + 90, 100, 200, 25))

        # 位置 过滤规则
        self.label_display_filter.setGeometry(QtCore.QRect(int(change * 40), 170, 90, 25))
        self.lineEdit_display_filter.setGeometry(QtCore.QRect(int(change * 40) + 100, 170, 200, 25))

        # 位置 确认按钮
        self.button_begin.setGeometry(QtCore.QRect(int(change * 400) + 20, 155, 120, 45))
        self.button_test.setGeometry(QtCore.QRect(int(change * 400) + 180, 155, 120, 45))

        # 显示区域
        self.tableWidget_overall.setGeometry(QtCore.QRect(40, 210, self.window_width - 85, 85))
        self.tableWidget.setGeometry(QtCore.QRect(40, 305, self.window_width - 85, self.window_height - 350))

        # 导出数据
        self.button_save_excel.setGeometry(QtCore.QRect(self.window_width - 100, self.window_height - 40, 80, 30))

    def choose_tshark(self):
        """
        获得 tshark 的文件位置
        :return: None
        """
        folder_path = QFileDialog.getOpenFileName(self, "选择tshark路径")[0]
        self.lineEdit_tshark.clear()
        self.lineEdit_tshark.setText(folder_path)

    def choose_live_capture(self):
        """
        选择活动接口读取
        :return:
        """
        self.choice_function = "live_capture"

    def choose_file_capture(self):
        """
        获得要读取 capture 文件位置
        :return:
        """
        folder_path = QFileDialog.getOpenFileName(self, "选择pcap路径")[0]
        self.lineEdit_file_capture.clear()
        self.lineEdit_file_capture.setText(folder_path)
        self.choice_function = "file_capture"

    def begin(self):
        self.button_begin.setEnabled(False)  # 设置按钮失效
        self.button_begin.setStyleSheet('''QPushButton{background:#D3D3D3}''')
        time.sleep(0.1)
        self.tshark_path = self.lineEdit_tshark.text()
        self.display_filter = self.lineEdit_display_filter.text()
        self.interface = self.comboBox_live_interface.currentText()
        self.file_rode = self.lineEdit_file_capture.text()
        if self.shark is not None:
            self.shark.close()
        self.shark = shark.WiresharkAnalysis(tshark_path=self.tshark_path, interface=None, bpf_filter=None,
                                             keep_packets=True, display_filter=self.display_filter)
        if self.choice_function == "live_capture":
            self.shark.pyshark_live_capture(display_filter_=self.display_filter)
        elif self.choice_function == "file_capture":
            self.shark.pyshark_file_capture(file_rode=self.file_rode, display_filter_=self.display_filter)
            ja3_dict = self.shark.count_file_ja3()
            self.repaint_table(ja3_dict)
            infor = self.shark.information_()
            self.repaint_table_overall(infor)
        else:
            pass
        self.button_begin.setStyleSheet('''QPushButton{background:#FFFFFF}''')
        self.button_begin.setEnabled(True)

    def repaint_table(self, ja3_dict):
        """
        输出获得的 TLS 指纹
        :param ja3_dict: dict
        :return: None
        """
        self.tableWidget.clear()
        self.tableWidget.setRowCount(len(ja3_dict))
        self.tableWidget.setColumnCount(6)
        self.tableWidget.horizontalHeader().setSectionResizeMode(QHeaderView.ResizeToContents)
        num_row = 0
        for item in ja3_dict:
            num_col = 0
            new_item = QTableWidgetItem(str(ja3_dict[item]))
            self.tableWidget.setItem(num_row, num_col, new_item)
            list_tls = item.split(",")
            for j in list_tls:
                num_col += 1
                new_item = QTableWidgetItem(j)
                self.tableWidget.setItem(num_row, num_col, new_item)
            num_row += 1

    def repaint_table_overall(self, info):
        self.tableWidget_overall.clear()
        self.tableWidget_overall.setRowCount(1)
        self.tableWidget_overall.setColumnCount(len(info))
        num_col = 0
        self.tableWidget_overall.setHorizontalHeaderLabels(info.keys())
        self.tableWidget_overall.horizontalHeader().setSectionResizeMode(QHeaderView.Stretch)
        for item in info:
            new_item = QTableWidgetItem(str(info[item]))
            self.tableWidget_overall.setItem(0, num_col, new_item)
            num_col += 1

    def save_excel(self):
        """
        保存全部情况 和 TLS指纹
        :return: None
        """
        workbook = openpyxl.Workbook()
        over_all_sheet = workbook.active
        over_all_sheet.title = "OverAllSheet"
        for col in range(1, self.tableWidget_overall.columnCount() + 1):
            over_all_sheet.cell(1, col, self.tableWidget_overall.horizontalHeaderItem(col - 1).text())
        for col in range(1, self.tableWidget_overall.columnCount() + 1):
            over_all_sheet.cell(2, col, self.tableWidget_overall.item(0, col - 1).text())

        tls_sheet = workbook.create_sheet("TLS Sheet")
        for row in range(1, self.tableWidget.rowCount() + 1):
            for col in range(1, self.tableWidget.columnCount() + 1):
                tls_sheet.cell(row, col, self.tableWidget.item(row - 1, col - 1).text())
        try:
            file_name, file_type = QFileDialog.getSaveFileName(self, 'save file', './', "Excel files(*.xlsx)")
            workbook.save(file_name)
            QMessageBox.about(self, "提示", "保存成功!     ")
        except IOError:
            QMessageBox.about(self, "提示", "保存失败!     ")

    def test_reprint(self):
        """
        测试函数 ----- 功能测试
        :return: None
        """
        dic = {
            "771,4865-4866-4867-49195-49196-52393-49199-49200-52392-49161-49162-49171-49172-156-157-47-53,"
            "0-23-65281-10-11-35-16-5-13-51-45-43-21,29-23-24,0": 10}
        self.tableWidget.clear()
        self.tableWidget.setRowCount(1)
        self.tableWidget.setColumnCount(6)
        num_row = 0
        num_col = 0
        self.tableWidget.horizontalHeader().setSectionResizeMode(QHeaderView.ResizeToContents)
        self.tableWidget.setColumnWidth(0, 30)
        self.tableWidget.setColumnWidth(1, 40)
        # self.tableWidget.setColumnWidth(2, 230)
        # self.tableWidget.setColumnWidth(3, 230)
        self.tableWidget.setColumnWidth(4, 120)
        self.tableWidget.setColumnWidth(5, 25)
        for item in dic:
            num_col = 0
            new_item = QTableWidgetItem(str(dic[item]))
            self.tableWidget.setItem(num_row, num_col, new_item)
            list_tls = item.split(",")
            for j in list_tls:
                num_col += 1
                new_item = QTableWidgetItem(j)
                self.tableWidget.setItem(num_row, num_col, new_item)
            num_row += 1
        info = {"TLS count": 100, "TCP count": 10, "all count": 200, "all Length": 78833, "TLS Length": 70000}

        self.tableWidget_overall.clear()
        self.tableWidget_overall.setRowCount(1)
        self.tableWidget_overall.setColumnCount(len(info))

        num_col = 0
        self.tableWidget_overall.setHorizontalHeaderLabels(info.keys())
        self.tableWidget_overall.horizontalHeader().setSectionResizeMode(QHeaderView.Stretch)

        for item in info:
            new_item = QTableWidgetItem(str(info[item]))
            self.tableWidget_overall.setItem(0, num_col, new_item)
            num_col += 1


if __name__ == '__main__':
    app = QtWidgets.QApplication(sys.argv)

    Main = MainUI()
    Main.show()

    sys.exit(app.exec_())
