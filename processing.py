import os
import sys

import openpyxl
from PyQt5.QtWidgets import *
from PyQt5 import QtCore, QtWidgets

from matplotlib import pyplot as plt


def print_bar(x, y, title, x_label, y_label):
    plt.bar(x, y, align='center')
    plt.title(title)
    plt.ylabel(y_label)
    plt.xlabel(x_label)
    plt.show()


class MainForm(QWidget):
    def __init__(self):
        super(MainForm, self).__init__()
        self.file = None
        self.setWindowTitle("处理窗口")
        self.resize(900, 600)  # 设置窗体大小

        # 选择文件夹，批量处理
        self.button_chooseDir = QPushButton(self)
        self.button_chooseDir.setText("选择文件夹")
        self.button_chooseDir.setGeometry(QtCore.QRect(20, 20, 100, 40))
        self.button_chooseDir.clicked.connect(self.get_excel_file_name)

        # 密码套件，批量处理
        self.button_Ciphers = QPushButton(self)
        self.button_Ciphers.setText("密码套件")
        self.button_Ciphers.setGeometry(QtCore.QRect(170, 50, 100, 40))
        self.button_Ciphers.clicked.connect(self.ciphers_get)

        # 确认，批量处理
        self.button_Extensions = QPushButton(self)
        self.button_Extensions.setText("扩展")
        self.button_Extensions.setGeometry(QtCore.QRect(170, 110, 100, 40))
        self.button_Extensions.clicked.connect(self.extensions_get)

        # TLS版本，批量处理
        self.button_Extensions = QPushButton(self)
        self.button_Extensions.setText("TLS版本")
        self.button_Extensions.setGeometry(QtCore.QRect(170, 170, 100, 40))
        self.button_Extensions.clicked.connect(self.version_get)

        # APP TLS分布，批量处理
        self.button_Extensions = QPushButton(self)
        self.button_Extensions.setText("APP TLS版本")
        self.button_Extensions.setGeometry(QtCore.QRect(170, 230, 110, 40))
        self.button_Extensions.clicked.connect(self.app_version_get)

        # 显示区  暂时关闭
        self.groupBox = QtWidgets.QGroupBox(self)
        self.groupBox.setGeometry(QtCore.QRect(300, 10, 580, 580))
        self.groupBox.setObjectName("groupBox")

    def get_excel_file_name(self):
        self.file = []
        path = QFileDialog.getExistingDirectory(self, 'choose dir', './')
        for root, dirs, files in os.walk(path):
            for file in files:
                if "~$" not in file:
                    self.file.append(os.path.join(root, file))
        print(self.file)

    def ciphers_get(self):
        data = self.count_num(3)
        x = data.keys()
        y = data.values()
        title = "Cipher Suite"
        x_label = "ciphers suite"
        y_label = "Count"
        print_bar(x, y, title, x_label, y_label)

    def extensions_get(self):
        data = self.count_num(4)
        x = data.keys()
        y = data.values()
        title = "Extensions"
        x_label = "extensions"
        y_label = "Count"
        print_bar(x, y, title, x_label, y_label)

    def version_get(self):
        version, proportion = self.count_version()
        print(version)
        print(proportion)
        sum_version = 0
        for i in version.values():
            sum_version += i
        frac = [i / sum_version for i in version.values()]
        explode = (0.05, 0.04, 0.03, 0.02)
        plt.title("TLS Version")
        patches, texts, auto_texts = plt.pie(frac, explode=explode, labels=version.keys(), autopct="%.2f%%",
                                             pctdistance=1.01, labeldistance=1.2)
        texts[0].set_y(-0.1)
        auto_texts[0].set_y(-0.1)

        texts[1].set_y(0.1)
        auto_texts[1].set_y(0.1)

        plt.show()

    def app_version_get(self):
        version, proportion = self.count_version()
        labels = ["APP" + str(i+1) for i in range(len(proportion))]
        x = range(len(labels))
        width = 0.35
        bottom_y = [0] * len(labels)
        sums = [sum(i) for i in proportion]
        barr = ["", "", "", ""]
        for i in range(4):
            y = [proportion[a][i] / sums[a] for a in range(len(labels))]
            barr[i] = plt.bar(x, y, width, bottom=bottom_y)
            bottom_y = [(a + b) for a, b in zip(y, bottom_y)]
        plt.xticks(x, labels, rotation=90)
        plt.legend(barr[::-1], reversed(["TLS1.0", "TLS1.1", "TLS1.2", "TLS1.3"]))
        plt.title('App TLS Version')
        plt.show()

    def count_version(self):
        """
        返回 tls版本情况
        :return: dir_version = {"TLS 1.0": 0, "TLS 1.1": 0, "TLS 1.2": 0, "TLS 1.3": 0}
         proportion = [[app1tls1.0, app1tls1.1....],[],]
        """
        dir_version = {"TLS 1.0": 0, "TLS 1.1": 0, "TLS 1.2": 0, "TLS 1.3": 0}
        proportion = []

        for file_name in self.file:
            workbook = openpyxl.load_workbook(file_name)
            tls_sheet = workbook['OverAllSheet']

            dir_version["TLS 1.0"] += int(tls_sheet.cell(1, 6).value)
            dir_version["TLS 1.1"] += int(tls_sheet.cell(1, 7).value)
            dir_version["TLS 1.2"] += int(tls_sheet.cell(1, 8).value)
            dir_version["TLS 1.3"] += int(tls_sheet.cell(1, 9).value)
            count = [int(tls_sheet.cell(1, 6).value), int(tls_sheet.cell(1, 7).value), int(tls_sheet.cell(1, 8).value),
                     int(tls_sheet.cell(1, 9).value)]
            proportion.append(count)

        return dir_version, proportion

    def count_num(self, pos):
        """
        在 excel 文件中读取数据
        :param pos: pos = 3 密码套件 ， pos = 4 扩展
        :return: {"编号":数量,"编号":数量}
        """
        dir_ans = {}
        for file_name in self.file:
            workbook = openpyxl.load_workbook(file_name)
            tls_sheet = workbook['TLS Sheet']
            list_sheet = []
            for row in tls_sheet.iter_rows():
                tmp = row[pos - 1].value.split("-")
                for i in tmp:
                    if i not in list_sheet:
                        list_sheet.append(i)
            for i in list_sheet:
                if i in dir_ans.keys():
                    dir_ans[i] += 1
                else:
                    dir_ans[i] = 1
            workbook.close()
        return dir_ans


if __name__ == "__main__":
    app = QtWidgets.QApplication(sys.argv)

    Main = MainForm()
    Main.show()

    sys.exit(app.exec_())
